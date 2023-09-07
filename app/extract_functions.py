import PyPDF2
import os
from openpyxl import load_workbook


def main(pdf_paths, excel_file):
    # after selecting the pdf files and excel file, run this function to go over each pdf file and extract the data
    counter = 0
    for pdf_path in pdf_paths:
        pdf_data = extract_pdf_fields(pdf_path)
        add_data_to_excel(excel_file, pdf_data)
        counter += 1
    print(str(counter) + " pdf files were processed. ")
    print("Data added to Excel file. check the excel file in the output folder.")


def extract_pdf_fields(pdf_path):
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        fields = reader.get_fields()
        data = {}
        data['Married?'] = ''

        for key, value in fields.items() or key in data:
            print(key, value.value)
            if value.value == None and key != 'Married?' and key != 'single' :
                data[key] = "N/A"
            elif key == 'married':
                if value.value == '/married':
                    data['Married?'] = 'Yes'
            elif key == 'single':
                if value.value == '/single':
                    data['Married?'] = 'No'
                
            else:
                data[key] = value.value

            # this to make sure that the data is in the right order when it is added to the excel file
            new_order = ['Full_name', 'Date_of_birth', 'Address','City', 'State', 'Zip',
                        'Cell_phone','Home_phone', 'Married?', 'Spouse_name','Em_name','Em_relationship']

        new_dict = {k: data[k] for k in new_order}
        return new_dict


def add_data_to_excel(output_file, data):
    if os.path.exists(output_file):
        workbook = load_workbook(output_file)

        sheet_name = "Sheet1"  # Specify the sheet name as "Sheet 1"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        empty_row = 1
        while sheet.cell(row=empty_row, column=1).value is not None:
            empty_row += 1

        for col_idx, key in enumerate(data.keys(), start=1):
            sheet.cell(row=empty_row, column=col_idx, value=data[key])

        workbook.save(output_file)
        return True

